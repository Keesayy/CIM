# By Arthur Péraud 12/2025
import os, openpyxl, sys, subprocess, csv, io 
import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path


def Ltoi(c: str) -> int:
    """ """
    return ord(c) - ord("A")


def Find_sheet_index(wb : Workbook, sheet_name : str) -> int:
    """ """
    return wb.sheetnames.index(sheet_name) if sheet_name in wb.sheetnames else -1


def Read_mdb_table(path_mdb: str, table_name: str) -> pd.DataFrame:
    """ """
    cmd = ["mdb-export", path_mdb, table_name]
    result = subprocess.run(cmd, check=True, capture_output=True, text=True)
    data = io.StringIO(result.stdout)
    return pd.read_csv(data)


def Get_col_from_mdb(mdb_file : Path, table_name : str, col_name : str) -> np.array:
    """ """
    if not mdb_file.exists(): 
        raise FileNotFoundError(f"Fichier MDB manquant: {mdb_file}")

    df = Read_mdb_table(str(mdb_file), table_name)

    if col_name not in df.columns: 
        raise KeyError(f"La colonne 'Bin1Amptd' est introuvable dans {mdb_file} (table {table_name}).")
    df = df[col_name]

    return df.to_numpy()


def Get_unique_filename(path : str) -> str:
    """Génère un nom de fichier unique en ajoutant (1), (2), etc. s'il existe déjà."""
    base, ext = os.path.splitext(path)
    i = 1
    new_path = path
    while os.path.exists(new_path):
        new_path = f"{base}({i}){ext}"
        i += 1
    return new_path


def Save_workbook_safely(wb : Workbook, output_file: str) -> None:
    """Sauvegarde fichier excel, confirmation si existence sinon préfixe est ajouté (Get_unique_filename)."""
    if os.path.exists(output_file):
        confirm = input(f"⚠️  Le fichier '{output_file}' existe déjà. Voulez-vous l’écraser ? (o/n) : ").strip().lower()
        
        if confirm == 'o' or confirm == 'y':
            wb.save(output_file)
            print(f"✅ Fichier écrasé et sauvegardé sous {output_file}")
        elif confirm == 'n':
            new_file = Get_unique_filename(output_file)
            wb.save(new_file)
            print(f"📁 Fichier sauvegardé sous un nouveau nom : {new_file}")
        else:
            print("Réponse non reconnue")
    else:
        wb.save(output_file)
        print(f"✅ Fichier sauvegardé sous {output_file}")


def Add_offset(arr: np.ndarray, spacing: float = 0.1, min_range: float = 1.0) -> np.ndarray:
    """"""
    n_traces, _ = arr.shape

    # Center each trace (preserves shape, avoids drift)
    centered = arr - arr.mean(axis=1, keepdims=True)

    # Vertical offsets
    offsets = np.arange(n_traces)[:, None] * spacing

    stacked = centered + offsets

    # Enforce minimum total range
    y_min, y_max = stacked.min(), stacked.max()
    current_range = y_max - y_min

    if current_range < min_range:
        stacked += (min_range - current_range) / 2

    return -stacked

def Fill_sheet_from_channel_prn(ws : Worksheet, data_path1 : Path, channel : str, sxx : str, letter : str, start_row : int) -> None:
    prn_file = data_path1 / f"Ch{channel}" / f"{channel}_{sxx}.prn"
    if not prn_file.exists(): raise FileNotFoundError(f"Missing file: {prn_file}")
    row = start_row

    with prn_file.open("r") as f:
        next(f)  # "S21 Log Mag"
        next(f)  # "Frequency (Hz)","dB",

        for line in f:
            line = line.strip()
            if not line: continue

            parts = [p.strip() for p in line.split(",") if p.strip()]
            db_value = round(float(parts[1]), 3)

            ws[f"{letter}{row}"] = db_value
            ws[f"{letter}{row}"].number_format = "0.000"
            row += 1

def Fill_sheet_from_channel(
            ws : Worksheet, 
            data_path1 : Path, 
            data_path2 : Path, 
            arr1 : np.ndarray, 
            arr2 : np.ndarray,             
            arr1_offset : np.ndarray, 
            arr2_offset : np.ndarray, 
            channel : str, 
            start_row : int = 3) -> None:
    """ """
    # PRN 
    Fill_sheet_from_channel_prn(ws, data_path1, channel, "S21", "E", start_row)
    Fill_sheet_from_channel_prn(ws, data_path1, channel, "S22", "AU", start_row)
    Fill_sheet_from_channel_prn(ws, data_path1, channel, "S11", "AV", start_row)

    # MDB
    all_channels = "ABCDEFGHIJKLMNOP" 
    if channel not in all_channels : raise ValueError(f"Canal invalide: {channel}")
    
    band_str = data_path1.name[6:10]
    if band_str == "0120" : band_str = "218"

    seq_without_channel = "".join(c for c in all_channels if c != channel)
    mdb_name = f"SP16T{band_str}_{seq_without_channel}_0.MDB"
    mdb_file = data_path2 / mdb_name

    arr = Get_col_from_mdb(mdb_file, "RasterScan", "Bin1Amptd")
    data_mdb = arr.reshape(-1,16)

    columns = [chr(c) for c in range(ord("X"), ord("Z") + 1)] + ["A" + chr(c) for c in range(ord("A"), ord("M") + 1)]

    row = start_row
    for i in range(data_mdb.shape[0]):
        for j, col in enumerate(columns):
            ws[f"{col}{row}"] = float(data_mdb[i, j])
            ws[f"{col}{row}"].number_format = "0.000"
        row += 1

    # Dynamic MDB
    k = Ltoi(channel)    
    for i in range(len(arr1[k])):
        ws[f"AP{i + start_row}"] = float(arr1_offset[k, i])
        ws[f"AP{i + start_row}"].number_format = "0.000"        
        ws[f"AQ{i + start_row}"] = float(arr2_offset[k, i])
        ws[f"AQ{i + start_row}"].number_format = "0.000"        

        ws[f"AR{i + start_row}"] = float(arr1[k, i])
        ws[f"AR{i + start_row}"].number_format = "0.000"        
        ws[f"AS{i + start_row}"] = float(arr2[k, i])
        ws[f"AS{i + start_row}"].number_format = "0.000"


def Fill_voies_sheets(input_file : Path, data_path1 : Path, data_path2 : Path, log_func = print) -> Workbook:
    wb = openpyxl.load_workbook(input_file)
    index = Find_sheet_index(wb, "Voie A")
    
    if index == -1: raise ValueError("La feuille 'Voie A' est introuvable dans le classeur.")
    wb.active = index

    # Channels from A to P
    channels = [f"Voie {chr(c)}" for c in range(ord("A"), ord("P") + 1)]

    # Dynamic arr
    band_str = data_path1.name[6:10]
    if band_str == "1840" :
        tmp1 = Get_col_from_mdb(data_path2 / "SP16TDynamic_20GHz_Even_0.MDB", "RasterScan", "Bin1Amptd")
        tmp2 = Get_col_from_mdb(data_path2 / "SP16TDynamic_20GHz_Even_1.MDB", "RasterScan", "Bin1Amptd")
        tmp3 = Get_col_from_mdb(data_path2 / "SP16TDynamic_20GHz_Odd_0.MDB", "RasterScan", "Bin1Amptd")
        tmp4 = Get_col_from_mdb(data_path2 / "SP16TDynamic_20GHz_Odd_1.MDB", "RasterScan", "Bin1Amptd")

        n = len(tmp1)
        arr1 = np.empty(2 * n, dtype=tmp1.dtype)
        arr2 = np.empty(2 * n, dtype=tmp1.dtype)
        
        arr1[0::2] = tmp1 
        arr1[1::2] = tmp3 
        arr2[0::2] = tmp2 
        arr2[1::2] = tmp4 

        arr1 = arr1.reshape(-1, 16).T
        arr2 = arr2.reshape(-1, 16).T
     
    else:
        tmp1 = Get_col_from_mdb(data_path2 / "SP16TDynamic_5GHz_1to8_0.MDB", "RasterScan", "Bin1Amptd").reshape(-1, 8)
        tmp2 = Get_col_from_mdb(data_path2 / "SP16TDynamic_5GHz_1to8_1.MDB", "RasterScan", "Bin1Amptd").reshape(-1, 8)
        tmp3 = Get_col_from_mdb(data_path2 / "SP16TDynamic_5GHz_9toF_0.MDB", "RasterScan", "Bin1Amptd").reshape(-1, 8)
        tmp4 = Get_col_from_mdb(data_path2 / "SP16TDynamic_5GHz_9toF_1.MDB", "RasterScan", "Bin1Amptd").reshape(-1, 8)

        arr1 = np.concatenate((tmp1, tmp3), axis = 1).T
        arr2 = np.concatenate((tmp2, tmp4), axis = 1).T

    arr1_offset = Add_offset(arr1)    
    arr2_offset = Add_offset(arr2)    

    k = 0
    for ch in channels:
        if ch in wb.sheetnames:
            ws = wb[ch]
            log_func(f"Remplissage de la sheet 'Voie {ch[5]}'")
            Fill_sheet_from_channel(ws, data_path1, data_path2, arr1, arr2, arr1_offset, arr2_offset, ch[5])
        else:
            raise ValueError(f"Sheet '{ch}' does not exist in the workbook")
    
    return wb


def Build_path_names(client: str, year: int, freqband: str, sn: str) -> tuple[Path, Path, Path, Path]:
    # Vérif bande de fréquence
    if len(freqband) != 4 or not freqband.isdigit(): raise ValueError("freqband doit être 4 chiffres, ex : '0120'.")

    # Vérif numéro de série
    if len(sn) != 4 or not sn.isdigit():  raise ValueError("sn doit être 4 chiffres, ex : '1910'.")

    A_str = freqband[:2]
    B_str = freqband[2:]

    if A_str[0] == "0": A_aff = "." + A_str[1]
    else: A_aff = A_str    

    band_str = f"{A_aff}-{B_str}"

    tmp1 = f"SP16T-{freqband}_Cal-E8361A"
    tmp1bis = f"SP16T-{freqband}_Iso_Dynamic"
    tmp2 = f"SP16T {band_str}GHz SN{sn} {year}.xlsx"
    tmp3 = f"SP16T {band_str}GHz SN{sn} 20XX.xlsx" # Fichier modèle pour le sheet modèle

    base_dir = Path(rf"E:\\Cal Info Mesure\\{client}\\Data {year}")
    
    data_path1 = base_dir / tmp1
    data_path2 = base_dir / tmp1 / tmp1bis
    output_file = base_dir / tmp2
    input_file = Path(tmp3)

    return data_path1, data_path2, output_file, input_file


### Main Program
if __name__ == "__main__":
    print("Bienvenue dans le programme Cal Info Mesure de SWITCH SPXT.")
    try:
        year = int(input("Entrez l'année : "))
        client = input("Entrez nom du client : ")
        freqband = input("Entrez la bande de fréquence (XXXX) : ")
        sn = input("Entrez le numéro de série (XXXX) : ")

        input_file = Path("SP16T .1-20GHz SN1910 20XX.xlsx") 
        data_path1 = Path("./SP16T-0120_Cal-E8361A")
        data_path2 = Path("./SP16T-0120_Iso_Dynamic")
        output_file = Path(f"SP16T .1-20GHz SN1910 {year}.xlsx")
        # data_path1, data_path2, output_file, input_file = Build_path_names(client, year, freqband, sn)

        wb = Fill_voies_sheets(input_file, data_path1, data_path2)
        Save_workbook_safely(wb, output_file)

    except Exception as e:
        print(f"Une erreur s'est produite : {e}")
